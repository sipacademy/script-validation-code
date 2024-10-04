import openpyxl
import os

# List of file paths (replace with your actual file paths)
file_paths = [
   # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Certificate.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Attendance.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Batch.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\School.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Student.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\CourseIntructor.xlsx',
   # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Dispatch Redeem product.xlsx',
    #'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Franchisee.xlsx',
    # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\InternalPurchaseOrder.xlsx',
    # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Invoice.xlsx',
    # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Invoice_item.xlsx',
    'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\LevelDiscontinue.xlsx',
    # 'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\PaymentReceipt.xlsx',
     'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Promocode.xlsx',
     'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\PurchaseReturn.xlsx',
     'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\Purchase_Redeem_point.xlsx',
     'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\PurchaseOrder.xlsx',
     'C:\\Users\\manmohan.d\\OneDrive - SAKSOFT LIMITED\\Desktop\\backup\\Subash\\executeKL\\SMSOrder.xlsx',
]

# Columns to check for empty values and duplicates for each sheet
sheets_columns_to_check = {
    # 'Certificate': {
    #     'empty_check': ('B','C','D','E','I','J','K','L','M','N','O','P','Q','R','AL','AM'),
    #     #'duplicate_check': ('B', 'C', 'D', 'N')
    #     'duplicate_check': ()  # No duplicate check for this sheet
    # },
    # 'Attendance': {
    #     'empty_check': ('A','B','C','D','E','F','G','H','I'),
    #     'duplicate_check': ()  # No duplicate check for this sheet
    # },
#      'Batch': {
#         'empty_check': ('B','C','D','E','F','G'),
#         'duplicate_check': ('B')
#    },
#     'School': {
#         'empty_check': ('B','C','D'),
#         'duplicate_check': ('B','C', 'D')
#     },
#     'CourseIntructor': {
#         'empty_check': ('B','C','F','G','H','J'),
#         'duplicate_check': ('B', 'C','H')
#     },
#     'Student': {
#         'empty_check': ('B','C','D','F','G','J','L','M','N','O','Q','R','T','V','X','Y','Z','AB','AC','AE','AH','AJ','AK','AL','AM','AN','AO','AP','AQ'),
#         'duplicate_check': ('F', 'N', 'Q', 'AH')
#     },
#     'Dispatch Redeem product': {
#         'empty_check': ('A','B','C','D','E','F','Q','R','S'),
#         'duplicate_check': ()  # No duplicate check for this sheet
#     },
    #  'Franchisee': {
    #      'empty_check': ('B','C','E','F','H','J','K','L','M','N','O','U','V','W','X','Y','AA','AC','AD','AG','AU','AX','AY','BA','BB','BC','BD','BE','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT'),
    #      'duplicate_check': ('B','H','U' 'AC','AD','',)
    #  },
   
    # 'InternalPurchaseOrder': {
    #     'empty_check': ('A','B','C','E','G','H','I','O','P'),
    #     'duplicate_check': ()  # No duplicate check for this sheet
    # },
    'Invoice': {
        'empty_check': ('A','B','C','D','E','F','G','H','I','J','L','M','P','W','Y','AB','AC','AD','AE','AF','AH','AK'),
        'duplicate_check': ('A', 'B', 'C', 'D', 'E')
    },
    'Invoice_item': {
        'empty_check': ('A','B','C','D','E','G','H','I','J','K','L'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'LevelDiscontinue': {
        'empty_check': ('B','C','D','E','H','I','J','L','M','N','O','P'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'Promocode': {
        'empty_check': ('B','D','F','G','H','I','K','L','Q','S'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'PurchaseReturn': {
        'empty_check': ('A','B','C','D','E','F','G','H','I','L','M','O','P','Q','R','S','T'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'Purchase_Redeem_point': {
        'empty_check': ('A','B','C','D','E','F','G','H','I','J','K','L'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'PurchaseOrder': {
        'empty_check': ('A','B','C','D','E','G','H','I','J','K','L','M','O','P','Q','U','X','Y','Z','AA'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    'SMSOrder': {
        'empty_check': ('A','B','C','D','E'),
        'duplicate_check': ()  # No duplicate check for this sheet
    },
    # Add more as needed
}

def process_sheet(sheet, sheet_name, columns_to_check):
    # Add a "Description" column header to the existing sheet if it doesn't exist
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Description":
            sheet.delete_cols(col)
            break

    description_col_idx = sheet.max_column + 1
    if sheet.cell(row=1, column=description_col_idx).value != "Description":
        sheet.cell(row=1, column=description_col_idx).value = "Description"

    # Dictionary to store combinations of columns to check for duplicates
    seen_combinations = {}

    # Loop through rows and check for empty values and duplicates
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        empty_check_cols = columns_to_check['empty_check']
        duplicate_check_cols = columns_to_check.get('duplicate_check')

        # Fetch values for empty check columns
        empty_values = [sheet[f'{col}{row[0].row}'].value for col in empty_check_cols]

        is_empty = False
        description = ""
        empty_columns = []

        # Clear the existing description for the row
        sheet.cell(row=row[0].row, column=description_col_idx).value = ""

        # Check for empty values in the specified columns and store empty columns
        for col, val in zip(empty_check_cols, empty_values):
            if val is None or str(val).strip() == "":
                empty_columns.append(col)

        # If any columns are empty, add them to the description
        if empty_columns:
            is_empty = True
            description += f"Empty values found in columns: {', '.join(empty_columns)} "
        # Specific checks for the "Student" sheet
        if sheet_name == "InternalPurchaseOrder":
            # Student validation code here...
             # Validate column B
            value_b = sheet[f'B{row[0].row}'].value
            if value_b not in ["SIP ABACUS", "SIP ABACUS INTERNATIONAL"]:
                description += f"Wrong value in column B: {value_b}. "

        # Specific checks for the "Student" sheet
        if sheet_name == "Student":
            # Student validation code here...
             # Validate column B
            value_b = sheet[f'B{row[0].row}'].value
            if value_b not in ["SIP ABACUS", "SIP ABACUS INTERNATIONAL"]:
                description += f"Wrong value in column B: {value_b}. "

            # Validate column J
            value_j = sheet[f'J{row[0].row}'].value
            if value_j not in ["New", "Transfer"]:
                description += f"Wrong value in column J: {value_j}. "

            # Validate column V
            value_v = sheet[f'V{row[0].row}'].value
            valid_v_values = [
                "Junior Level 1", "Junior Level 2", "Junior Level 3", "Junior Level 4",
                "Foundation Level 1", "Foundation Level 2", "Foundation Level 3", "Foundation Level 4",
                "Grand Module A", "Grand Module B", "Grand Module C",
                "Advance Level 1", "Advance Level 2", "Advance Level 3", "Advance Level 4"
            ]
            if value_v not in valid_v_values:
                description += f"Wrong value in column V: {value_v}. "

            # Validate column Z
            value_z = sheet[f'Z{row[0].row}'].value
            if value_z not in ["Active", "Discontinued", "Complete"]:
                description += f"Wrong value in column Z: {value_z}. "

            # Validate column AL
            value_al = sheet[f'AL{row[0].row}'].value
            if value_al not in valid_v_values:
                description += f"Wrong value in column AL: {value_al}. "

            # Validate column AO
            value_ao = sheet[f'AO{row[0].row}'].value
            if value_ao not in ["SIP ABACUS", "SIP ABACUS INTERNATIONAL"]:
                description += f"Wrong value in column AO: {value_ao}. "

        # Add validation for the "Certificate" sheet
        if sheet_name == "Certificate":
            # Validate Column I (non-case-sensitive)
            value_i = str(sheet[f'I{row[0].row}'].value).strip().lower() if sheet[f'I{row[0].row}'].value else ""
            if value_i not in ["sip abacus", "sip abacus international"]:
                description += f"Wrong value in column I: {sheet[f'I{row[0].row}'].value}. "

            # Validate Column J (non-case-sensitive)
            value_j = str(sheet[f'J{row[0].row}'].value).strip().lower() if sheet[f'J{row[0].row}'].value else ""
            valid_j_values = [
                "junior level 1", "junior level 2", "junior level 3", "junior level 4",
                "foundation level 1", "foundation level 2", "foundation level 3", "foundation level 4",
                "grand module a", "grand module b", "grand module c",
                "advance level 1", "advance level 2", "advance level 3", "advance level 4"
            ]
            if value_j not in valid_j_values:
                description += f"Wrong value in column J: {sheet[f'J{row[0].row}'].value}. "

            # Validate Column L (non-case-sensitive)
            value_l = str(sheet[f'L{row[0].row}'].value).strip().lower() if sheet[f'L{row[0].row}'].value else ""
            if value_l not in ["active", "discontinued", "complete"]:
                description += f"Wrong value in column L: {sheet[f'L{row[0].row}'].value}. "

            # Validate Column M (non-case-sensitive)
            value_m = str(sheet[f'M{row[0].row}'].value).strip().lower() if sheet[f'M{row[0].row}'].value else ""
            valid_m_values = [
                "not yet booked", "booked", "requisition", "approved", "dispatched",
                "partial requisition", "received", "reassessment"
            ]
            if value_m not in valid_m_values:
                description += f"Wrong value in column M: {sheet[f'M{row[0].row}'].value}. "

            # Validate Column AL (non-case-sensitive)
            value_al = str(sheet[f'AL{row[0].row}'].value).strip().lower() if sheet[f'AL{row[0].row}'].value else ""
            if value_al not in valid_m_values:
                description += f"Wrong value in column AL: {sheet[f'AL{row[0].row}'].value}. "
       
        # Check for duplicates only if `duplicate_check_cols` is provided
        if duplicate_check_cols and not is_empty:
            dup_values = tuple(sheet[f'{col}{row[0].row}'].value for col in duplicate_check_cols)

            # If the combination is already in seen_combinations, mark it as a duplicate
            if dup_values in seen_combinations:
                description += f"Duplicate values found in columns {', '.join(duplicate_check_cols)} "
            else:
                # Store this combination as seen with the current row number
                seen_combinations[dup_values] = row[0].row

        # Write the new description if the row has any empty columns or duplicates
        if description:
            sheet.cell(row=row[0].row, column=description_col_idx).value = description.strip()
        else:
            sheet.cell(row=row[0].row, column=description_col_idx).value = "value is correct"  # If all checks passed

def process_file(file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        print(f"Processing file: {file_path}")

        # Check each sheet in the file based on the file name
        file_name = os.path.basename(file_path).replace('.xlsx', '')
        if file_name in sheets_columns_to_check:
            sheet_name = file_name
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                process_sheet(sheet, sheet_name, sheets_columns_to_check[sheet_name])
            else:
                print(f"Sheet '{sheet_name}' not found in {file_path}")

        # Save the workbook back with the results
        wb.save(file_path)
        print(f"Results saved for {file_path}")
    except Exception as e:
        print(f"An error occurred with file '{file_path}': {e}")

# Process all files in the list
for file_path in file_paths:
    process_file(file_path)

print("All files processed successfully.")
